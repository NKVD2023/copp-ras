import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1, value="Very very very long string without newlines to see what happens")
ws.column_dimensions['A'].width = 10
ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)

def _autosize_excel(ws, min_col_width=12, max_col_width=60, default_row_height=15):
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            if cell.value:
                print(f"Row {cell.row} Col {cell.column} value={cell.value}")
                try:
                    lines = str(cell.value).split("\n")
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
                except:
                    pass
        
        if col_letter:
            adjusted_width = min(max(max_length + 2, min_col_width), max_col_width)
            ws.column_dimensions[col_letter].width = adjusted_width

_autosize_excel(ws)
print("After autosize width is:", ws.column_dimensions['A'].width)

