"""
Модуль отчетов: Импорт данных и шаблонов (Reports - Import).
Предоставляет функционал:
1. Скачивание пустого Excel-бланка отчета (для заполнения оффлайн).
2. Загрузка (импорт) заполненного Excel-файла за выбранную организацию.
Парсер автоматически сопоставляет данные из файла с полями схемы.
"""
from flask import request, jsonify, send_file
from flask_login import login_required, current_user
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re
from app import db
from app.reports import reports_bp
from app.models import ReportTemplate, ReportSubmission, User

# ==========================================
# ЭКСПОРТ ПУСТОЙ ФОРМЫ И ИМПОРТ ИЗ EXCEL
# ==========================================

@reports_bp.route('/export_blank/<int:template_id>')
@login_required
def export_blank(template_id):
    """
    Генерация пустого Excel-шаблона для заполнения учреждениями вне системы (оффлайн).
    Формат генерируемого файла идентичен сводному экспорту (заголовок, шапка),
    но содержит только одну пустую строку для ввода данных.
    Это необходимо, чтобы парсер `import_excel` мог легко разобрать этот файл.
    """
    if current_user.role not in ['admin', 'viewer']:
        return "Доступ ограничен", 403

    template = ReportTemplate.query.get_or_404(template_id)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- НАСТРОЙКИ СТИЛЕЙ EXCEL (идентичны export_excel) ---
    title_font = Font(name='Arial', size=14, bold=True, color="000000")
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0071DC")
    data_font = Font(name='Arial', size=11)

    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- ГЕНЕРАЦИЯ ЛИСТОВ ИЗ СХЕМЫ ---
    for sheet_data in template.schema:
        safe_title = re.sub(r'[\\*?:/\[\]]', '', sheet_data['sheet_title'])[:31]
        ws = wb.create_sheet(title=safe_title)

        max_col = len(sheet_data['fields']) + 1

        # 1. ЗАГОЛОВОК (название отчета, объединение ячеек)
        title_cell = ws.cell(row=1, column=1, value=template.name)
        title_cell.font = title_font
        title_cell.alignment = align_center

        for col in range(1, max_col + 1):
            ws.cell(row=1, column=col).border = thin_border

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.row_dimensions[1].height = 100

        # 2. ШАПКА ТАБЛИЦЫ
        headers = ['Организация'] + [f['label'] for f in sheet_data['fields']]
        ws.append(headers)
        ws.row_dimensions[2].height = 100

        # Настройка ширины
        ws.column_dimensions['A'].width = 50
        for col_idx in range(2, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 45

        # Стилизация шапки
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        # 3. ПУСТАЯ СТРОКА ДЛЯ ЗАПОЛНЕНИЯ УЧРЕЖДЕНИЕМ
        empty_row = [''] * max_col
        ws.append(empty_row)
        ws.row_dimensions[3].height = 70

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = align_left if col_idx == 1 else align_center

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Форма_{template.short_name}.xlsx".replace(" ", "_")

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@reports_bp.route('/import_excel/<int:template_id>', methods=['POST'])
@login_required
def import_excel(template_id):
    """
    Импорт заполненного Excel-файла. Создает НОВУЮ запись (ReportSubmission)
    от лица указанного учреждения (организации).
    
    Алгоритм парсера:
    1. Ищет листы по названию (или по индексу, если название изменено).
    2. Сканирует строку 2 на наличие известных заголовков полей (сверяет с schema).
    3. Сканирует строки с данными (начиная со строки 3).
    4. Преобразует типы данных (числа/текст) и собирает итоговый JSON `data`.
    
    Ожидает multipart/form-data:
    - file: .xlsx файл
    - user_id: ID пользователя (организации), за которую загружаются данные.
    """
    if current_user.role not in ['admin', 'viewer']:
        return jsonify({'status': 'error', 'message': 'Доступ ограничен'}), 403

    template = ReportTemplate.query.get_or_404(template_id)

    # --- ВАЛИДАЦИЯ ВХОДНЫХ ДАННЫХ И ФАЙЛОВ ---
    user_id = request.form.get('user_id')
    if not user_id:
        return jsonify({'status': 'error', 'message': 'Не выбран пользователь (организация).'}), 400

    user = User.query.get(int(user_id))
    if not user:
        return jsonify({'status': 'error', 'message': 'Пользователь не найден.'}), 404

    file = request.files.get('file')
    if not file:
        return jsonify({'status': 'error', 'message': 'Файл не выбран.'}), 400

    if not file.filename.endswith('.xlsx'):
        return jsonify({'status': 'error', 'message': 'Поддерживаются только файлы формата .xlsx'}), 400

    # --- ЧТЕНИЕ EXCEL ---
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Не удалось открыть файл: {str(e)}'}), 400

    data = {}  # Итоговый словарь данных {field_name: value}
    warnings = [] # Предупреждения (например, не найденные листы или заголовки)

    # --- ПАРСИНГ ДАННЫХ ПО ЛИСТАМ ---
    for sheet_idx, sheet_data in enumerate(template.schema):
        safe_title = re.sub(r'[\\*?:/\[\]]', '', sheet_data['sheet_title'])[:31]

        # Пытаемся найти лист по названию (приоритет), затем по индексу (fallback)
        ws = None
        if safe_title in wb.sheetnames:
            ws = wb[safe_title]
        elif sheet_idx < len(wb.sheetnames):
            ws = wb.worksheets[sheet_idx]
            warnings.append(f'Лист "{safe_title}" не найден, использован лист #{sheet_idx + 1} ("{wb.sheetnames[sheet_idx]}")')
        else:
            warnings.append(f'Лист "{safe_title}" не найден в файле — пропущен.')
            continue

        # Строим словарь соответствия: текст заголовка -> информация о поле (field dict)
        label_to_name = {}
        for field in sheet_data['fields']:
            label_to_name[field['label'].strip().lower()] = field

        # Находим строку с заголовками (обычно это строка 2 по формату пустого бланка)
        header_row = 2
        max_col = ws.max_column or 1

        # Составляем карту колонок: индекс колонки -> field_info
        col_map = {}
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row=header_row, column=col_idx).value
            if cell_value:
                header_text = str(cell_value).strip().lower()
                if header_text in label_to_name:
                    col_map[col_idx] = label_to_name[header_text]

        if not col_map:
            warnings.append(f'На листе "{safe_title}" не найдено совпадений заголовков — пропущен.')
            continue

        # Читаем строки данных (начиная с 3-й строки)
        for row_idx in range(3, (ws.max_row or 3) + 1):
            # Игнорируем строку "Итого", если она случайно попала в импорт
            first_cell = ws.cell(row=row_idx, column=1).value
            if first_cell and str(first_cell).strip().lower() == 'итого':
                continue

            # Проверяем, есть ли хоть какие-то непустые данные в этой строке
            row_has_data = False
            for col_idx, field_info in col_map.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip() != '':
                    row_has_data = True
                    break

            if not row_has_data:
                continue

            # Извлекаем и нормализуем данные ячеек
            for col_idx, field_info in col_map.items():
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    # Конвертация типов в зависимости от схемы
                    if field_info['type'] == 'number':
                        try:
                            val = float(cell_value)
                            # Если целое число — убираем суффикс .0 (12.0 -> "12")
                            if val == int(val):
                                data[field_info['name']] = str(int(val))
                            else:
                                data[field_info['name']] = str(val)
                        except (ValueError, TypeError):
                            # Если текст в числовом поле - оставляем как есть
                            data[field_info['name']] = str(cell_value).strip()
                    else:
                        data[field_info['name']] = str(cell_value).strip()
                else:
                    data[field_info['name']] = ''

    wb.close()

    # --- ПРОВЕРКА: собрали ли мы вообще данные? ---
    non_empty = {k: v for k, v in data.items() if v}
    if not non_empty:
        return jsonify({
            'status': 'error',
            'message': 'Файл не содержит данных для импорта (или заголовки не совпали).',
            'warnings': warnings
        }), 400

    # --- СОЗДАНИЕ НОВОГО ОТВЕТА (всегда новый объект, без обновления старого) ---
    submission = ReportSubmission(
        template_id=template.id,
        user_id=user.id,
        data=data
    )
    db.session.add(submission)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'message': f'Данные успешно импортированы для "{user.username}". Заполнено полей: {len(non_empty)}.',
        'imported_count': len(non_empty),
        'warnings': warnings
    })
