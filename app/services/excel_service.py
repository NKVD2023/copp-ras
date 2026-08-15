import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re

class ExcelService:
    @staticmethod
    def export_debtors(template, debtors):
        """
        Формирует Excel файл со списком должников.
        :param template: Объект ReportTemplate
        :param debtors: Список объектов User (должники)
        :return: (output: BytesIO, filename: str)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Должники"
        
        # Настройка заголовка
        ws.append(["Организация"])
        header_cell = ws.cell(row=1, column=1)
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="00334E", end_color="00334E", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_cell.border = thin_border
        
        # Ширина колонки
        ws.column_dimensions['A'].width = 80
        
        # Заполнение данных
        for row_idx, d in enumerate(debtors, start=2):
            org_name = d.description if d.description else d.username
            cell = ws.cell(row=row_idx, column=1, value=org_name)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Должники_{template.short_name}.xlsx".replace(" ", "_")
        return output, filename

    @staticmethod
    def export_statistics(stat_schema, short_name, user_title=None):
        """
        Формирует Excel файл со сводной статистикой.
        Макет: периоды вертикально (от новых к старым), поля горизонтально.
        :param stat_schema: Список листов (как в веб-дашборде)
        :param short_name: Тип отчета.
        :param user_title: Имя учреждения (для админа) или None (для пользователя).
        :return: (output: BytesIO, filename: str)
        """
        CORP_COLOR  = "00334E"
        COLOR_UP    = "16A34A"
        COLOR_DOWN  = "DC2626"
        COLOR_EMPTY = "94A3B8"

        title_font = Font(bold=True, size=16, color="000000", name='Arial')
        header_font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
        header_fill = PatternFill(start_color=CORP_COLOR, end_color=CORP_COLOR, fill_type="solid")
        period_font = Font(bold=True, color="1E293B", name='Arial', size=11)
        period_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        data_font   = Font(name='Arial', size=11)
        
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Удаляем дефолтный лист

        for sheet_index, sheet_data in enumerate(stat_schema):
            title = sheet_data.get('sheet_title', f'Лист {sheet_index+1}')
            safe_title = re.sub(r'[\\*?:/\[\]]', '', title)[:31]
            ws = wb.create_sheet(title=safe_title)

            header_rows = sheet_data.get('header_rows', [])
            leaf_fields = sheet_data.get('leaf_fields', [])
            periods_data = sheet_data.get('periods_data', [])
            
            max_depth = len(header_rows) if header_rows else 1
            max_col = len(leaf_fields) + 1

            # 1. ЗАГОЛОВОК ЛИСТА
            title_text = f"Статистика: {short_name}"
            if user_title:
                title_text += f" | {user_title}"
            
            title_cell = ws.cell(row=1, column=1, value=title_text)
            title_cell.font = title_font
            title_cell.alignment = align_center
            
            for col in range(1, max_col + 1):
                ws.cell(row=1, column=col).border = thin_border
                
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws.row_dimensions[1].height = 60

            # 2. РЕНДЕР СЛОЖНОЙ ШАПКИ
            start_row = 2
            
            for r in range(start_row, start_row + max_depth):
                ws.row_dimensions[r].height = 40
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = thin_border

            ws.cell(row=start_row, column=1, value="Период")
            if max_depth > 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + max_depth - 1, end_column=1)

            occupied = set()
            for row_idx, h_row in enumerate(header_rows):
                c_row = start_row + row_idx
                current_col = 2
                for cell_info in h_row:
                    while (c_row, current_col) in occupied:
                        current_col += 1
                        
                    ws.cell(row=c_row, column=current_col, value=cell_info['label'])
                    
                    for r in range(c_row, c_row + cell_info['rowspan']):
                        for c in range(current_col, current_col + cell_info['colspan']):
                            occupied.add((r, c))
                    
                    if cell_info['colspan'] > 1 or cell_info['rowspan'] > 1:
                        ws.merge_cells(
                            start_row=c_row, 
                            start_column=current_col, 
                            end_row=c_row + cell_info['rowspan'] - 1, 
                            end_column=current_col + cell_info['colspan'] - 1
                        )

            # 3. ДАННЫЕ
            current_row = start_row + max_depth
            
            for p_data in periods_data:
                period_label = p_data["period"]
                
                max_len = 1
                if p_data["has_submission"]:
                    for f in leaf_fields:
                        f_id = str(f.get('name') or f.get('id', ''))
                        val = p_data['values'].get(f_id)
                        if isinstance(val, list):
                            max_len = max(max_len, len(val))

                start_merge_row = current_row

                for i in range(max_len):
                    if i == 0:
                        period_cell = ws.cell(row=current_row, column=1, value=period_label)
                        period_cell.font = period_font
                        period_cell.fill = period_fill
                        period_cell.alignment = align_center
                    else:
                        period_cell = ws.cell(row=current_row, column=1)
                        
                    period_cell.border = thin_border

                    for j, f in enumerate(leaf_fields):
                        f_id = str(f.get('name') or f.get('id', ''))
                        cell = ws.cell(row=current_row, column=2 + j)
                        
                        cell_val = "-"
                        if p_data["has_submission"]:
                            val = p_data['values'].get(f_id)
                            if isinstance(val, list):
                                if i < len(val):
                                    cell_val = val[i]
                            else:
                                if i == 0:
                                    cell_val = val
                                else:
                                    cell_val = ""
                        else:
                            if i == 0:
                                cell_val = "Нет данных"
                            else:
                                cell_val = ""

                        if cell_val in [None, '', []]:
                            cell_val = "-"

                        is_numeric = str(f.get('type', '')).lower() in ['number', 'числовое']
                        if is_numeric and cell_val not in ["-", "Нет данных", ""]:
                            try:
                                cell_num = float(str(cell_val).replace(',', '.')) if "." in str(cell_val) or "," in str(cell_val) else int(cell_val)
                                cell_val = cell_num
                            except (ValueError, TypeError):
                                pass
                                
                        cell.value = cell_val if cell_val != "" else None
                        
                        if i == 0 and is_numeric and "deltas" in p_data and f_id in p_data["deltas"]:
                            delta = p_data["deltas"][f_id]
                            if delta > 0:
                                cell.font = Font(color=COLOR_UP, bold=True, name='Arial', size=11)
                            elif delta < 0:
                                cell.font = Font(color=COLOR_DOWN, bold=True, name='Arial', size=11)
                            else:
                                cell.font = data_font
                        elif cell_val == "Нет данных":
                            cell.font = Font(color=COLOR_EMPTY, italic=True, name='Arial', size=11)
                        else:
                            cell.font = data_font
                            
                        cell.alignment = align_center
                        cell.border = thin_border
                        
                    current_row += 1

                if max_len > 1:
                    ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=current_row-1, end_column=1)
                    
                    for j, f in enumerate(leaf_fields):
                        f_id = str(f.get('name') or f.get('id', ''))
                        val = p_data['values'].get(f_id) if p_data["has_submission"] else None
                        if not isinstance(val, list):
                            ws.merge_cells(start_row=start_merge_row, start_column=2+j, end_row=current_row-1, end_column=2+j)
                            cell.value = 0
                    else:
                        cell.value = "-"
                
                current_row += 1

        # Авто-размер ячеек
        ExcelService._autosize_excel(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = short_name.replace(" ", "_").replace("/", "-")
        filename = f"Статистика_{safe_name}.xlsx"
        return output, filename

    @staticmethod
    def export_report(template, submissions):
        """
        Формирует сводный Excel файл со сданными отчетами (поддержка сложной шапки).
        :param template: Объект ReportTemplate
        :param submissions: Список объектов ReportSubmission
        :return: (output: BytesIO, filename: str)
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # --- НАСТРОЙКИ СТИЛЕЙ ---
        title_font = Font(name='Arial', size=14, bold=True, color="000000")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="00334E")
        data_font = Font(name='Arial', size=11)
        total_font = Font(name='Arial', size=11, bold=True)
        total_fill = PatternFill("solid", fgColor="F8F9FA")
        
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for sheet_data in template.schema:
            safe_title = re.sub(r'[\\*?:/\[\]]', '', sheet_data['sheet_title'])[:31]
            ws = wb.create_sheet(title=safe_title)

            fields = sheet_data.get('fields', [])
            
            from app.utils import build_table_headers
            header_rows, leaf_fields = build_table_headers(fields)
            max_depth = len(header_rows) if header_rows else 1
            
            # Общее количество колонок = 1 (Организация) + количество полей данных
            max_col = len(leaf_fields) + 1

            # 1. ЗАГОЛОВОК ЛИСТА
            title_cell = ws.cell(row=1, column=1, value=template.name)
            title_cell.font = title_font
            title_cell.alignment = align_center
            
            for col in range(1, max_col + 1):
                ws.cell(row=1, column=col).border = thin_border
                
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws.row_dimensions[1].height = 60

            # 2. РЕНДЕР СЛОЖНОЙ ШАПКИ
            start_row = 2
            
            # Устанавливаем стили для всей области шапки заранее
            for r in range(start_row, start_row + max_depth):
                ws.row_dimensions[r].height = 40
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = thin_border

            # Заголовок первой колонки
            ws.cell(row=start_row, column=1, value="Организация")
            if max_depth > 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + max_depth - 1, end_column=1)

            # Отрисовываем древовидную шапку
            occupied = set()
            for row_idx, h_row in enumerate(header_rows):
                c_row = start_row + row_idx
                current_col = 2 # Начинаем со 2 колонки
                for cell_info in h_row:
                    # Пропускаем колонки, которые уже заняты через rowspan сверху
                    while (c_row, current_col) in occupied:
                        current_col += 1
                        
                    ws.cell(row=c_row, column=current_col, value=cell_info['label'])
                    
                    # Помечаем все ячейки, которые занимает этот заголовок, как занятые
                    for r in range(c_row, c_row + cell_info['rowspan']):
                        for c in range(current_col, current_col + cell_info['colspan']):
                            occupied.add((r, c))
                    
                    if cell_info['colspan'] > 1 or cell_info['rowspan'] > 1:
                        ws.merge_cells(
                            start_row=c_row, 
                            start_column=current_col, 
                            end_row=c_row + cell_info['rowspan'] - 1, 
                            end_column=current_col + cell_info['colspan'] - 1
                        )

            # Ширина колонок
            ws.column_dimensions['A'].width = 50 
            for col_idx in range(2, max_col + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 25

            # 3. ДАННЫЕ
            current_row = start_row + max_depth
            
            for sub in submissions:
                org_name = sub.user.description if sub.user.description else sub.user.username
                
                # 1. Determine if has_data and calculate max_len
                has_data = False
                max_len = 1
                for f in leaf_fields:
                    field_key = f.get('name') or f.get('id')
                    val = sub.data.get(field_key)
                    if val is not None:
                        if isinstance(val, list):
                            for v in val:
                                if v is not None and str(v).strip() != "":
                                    has_data = True
                            max_len = max(max_len, len(val))
                        elif str(val).strip() != "":
                            has_data = True
                
                if not has_data:
                    continue
                
                start_merge_row = current_row
                sub_totals = {f.get('name') or f.get('id'): 0.0 for f in leaf_fields if str(f.get('type', '')).lower() in ['number', 'числовое']}
                has_subtotals = {f.get('name') or f.get('id'): False for f in leaf_fields if str(f.get('type', '')).lower() in ['number', 'числовое']}

                # 2. Generate max_len rows
                for i in range(max_len):
                    row_data = [org_name if i == 0 else ""]
                    
                    for f in leaf_fields:
                        field_key = f.get('name') or f.get('id')
                        val = sub.data.get(field_key)
                        
                        f_type = str(f.get('type', '')).lower()
                        is_numeric = f_type in ['number', 'числовое']
                        
                        cell_val = "-"
                        if isinstance(val, list):
                            if i < len(val):
                                cell_val = val[i]
                        else:
                            if i == 0:
                                cell_val = val
                            else:
                                cell_val = ""
                                
                        if cell_val in [None, '', []]:
                            cell_val = "-"
                            
                        if is_numeric and cell_val != "-":
                            try:
                                cell_num = float(cell_val)
                                cell_val = cell_num
                                sub_totals[field_key] += cell_num
                                has_subtotals[field_key] = True
                            except (ValueError, TypeError):
                                pass
                                
                        row_data.append(cell_val)

                    ws.append(row_data)
                    
                    for col_idx, _ in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.font = data_font
                        cell.border = thin_border
                        cell.alignment = align_left if col_idx == 1 else align_center
                    
                    current_row += 1
                
                # 3. Merge vertical cells for scalars and org name
                if max_len > 1:
                    ws.merge_cells(start_row=start_merge_row, start_column=1, end_row=start_merge_row + max_len - 1, end_column=1)
                    
                    col_idx = 2
                    for f in leaf_fields:
                        field_key = f.get('name') or f.get('id')
                        val = sub.data.get(field_key)
                        if not isinstance(val, list):
                            ws.merge_cells(start_row=start_merge_row, start_column=col_idx, end_row=start_merge_row + max_len - 1, end_column=col_idx)
                        col_idx += 1
                        
                    # 4. Add subtotal row
                    subtotal_row = ['Итого по организации']
                    for f in leaf_fields:
                        field_key = f.get('name') or f.get('id')
                        f_type = str(f.get('type', '')).lower()
                        if f_type in ['number', 'числовое']:
                            if has_subtotals[field_key]:
                                subtotal_row.append(sub_totals[field_key])
                            else:
                                subtotal_row.append(0)
                        else:
                            subtotal_row.append('-')
                            
                    ws.append(subtotal_row)
                    
                    subtotal_font = Font(name='Arial', size=10, bold=True)
                    for col_idx, _ in enumerate(subtotal_row, 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.font = subtotal_font
                        cell.border = thin_border
                        cell.alignment = align_left if col_idx == 1 else align_center
                        # Highlight row with light gray background
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                            
                    current_row += 1

            # 4. ИТОГО
            total_row = ['Итого']
            for f in leaf_fields:
                f_type = str(f.get('type', '')).lower()
                if f_type in ['number', 'числовое']:
                    col_total = 0
                    has_data = False
                    for sub in submissions:
                        field_key = f.get('name') or f.get('id')
                        val = sub.data.get(field_key)
                        if val:
                            if isinstance(val, list):
                                for v in val:
                                    if v:
                                        try:
                                            col_total += float(v)
                                            has_data = True
                                        except (ValueError, TypeError):
                                            pass
                            else:
                                try:
                                    col_total += float(val)
                                    has_data = True
                                except (ValueError, TypeError):
                                    pass
                    total_row.append(col_total if has_data else 0)
                else:
                    total_row.append('-')

            ws.append(total_row)
            
            # Применяем авто-размер ячеек к текущему листу
            ExcelService._autosize_excel(ws)

            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = align_left if col_idx == 1 else align_center

        filename = f"Свод_{template.short_name}.xlsx".replace(" ", "_")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, filename

    @staticmethod
    def _autosize_excel(ws, min_col_width=12, max_col_width=60, default_row_height=15):
        from openpyxl.utils import get_column_letter
        
        # Авто-размер колонок
        for col in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                if col_letter is None:
                    col_letter = get_column_letter(cell.column)
                if cell.value:
                    try:
                        lines = str(cell.value).split("\n")
                        for line in lines:
                            if len(line) > max_length:
                                max_length = len(line)
                    except:
                        pass
            
            if col_letter:
                adjusted_width = min(max(max_length + 2, min_col_width), max_col_width)
                ws.column_dimensions[col_letter].width = adjusted_width
                
        # Авто-высота строк
        for row in ws.rows:
            if not row:
                continue
            max_lines = 1
            is_header = (row[0].row <= 2) # Добавим отступ для шапки
            
            for cell in row:
                if cell.value:
                    try:
                        col_letter = get_column_letter(cell.column)
                        col_width = ws.column_dimensions[col_letter].width
                        if not col_width:
                            col_width = min_col_width
                        
                        # Эффективная ширина символов (с учетом переносов слов и жирного шрифта)
                        effective_width = max(5, int(col_width) * 0.8)
                        
                        lines = str(cell.value).split("\n")
                        total_lines_for_cell = 0
                        for line in lines:
                            if len(line) == 0:
                                total_lines_for_cell += 1
                            else:
                                total_lines_for_cell += -(-len(line) // int(effective_width))
                                
                        if total_lines_for_cell > max_lines:
                            max_lines = total_lines_for_cell
                    except:
                        pass
                        
            # Если это первая строка (объединенная), она обычно длинная, но занимает много колонок
            if row[0].row == 1:
                ws.row_dimensions[row[0].row].height = 25
            else:
                padding = 15 if is_header else 5
                ws.row_dimensions[row[0].row].height = max_lines * default_row_height + padding
