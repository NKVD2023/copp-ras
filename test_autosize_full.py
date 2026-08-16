import json
import openpyxl
from openpyxl.utils import get_column_letter

def simulate():
    wb = openpyxl.Workbook()
    ws = wb.active

    # Use the schema snippet I dumped earlier
    header_rows = [
      [
        {
          "label": "Суммарный выпуск  2025 год, (человек)",
          "colspan": 2,
          "rowspan": 1
        }
      ],
      [
        {
          "label": "Всего",
          "colspan": 1,
          "rowspan": 3
        },
        {
          "label": "Трудоустроены (по трудовому договору, договору ГПХ в соответствии с трудовым законодательством, законодательством  об обязательном пенсионном страховании, чел.",
          "colspan": 1,
          "rowspan": 1
        }
      ],
      [
        {
          "label": "из них трудоустроенных по Республике Крым, чел.",
          "colspan": 1,
          "rowspan": 2
        }
      ],
      []
    ]
    
    start_row = 2
    occupied = set()
    for row_idx, h_row in enumerate(header_rows):
        c_row = start_row + row_idx
        current_col = 2
        for cell_info in h_row:
            while (c_row, current_col) in occupied:
                current_col += 1
                
            ws.cell(row=c_row, column=current_col, value=cell_info['label'])
            
            for r in range(c_row, c_row + cell_info.get('rowspan', 1)):
                for c in range(current_col, current_col + cell_info.get('colspan', 1)):
                    occupied.add((r, c))
            
            if cell_info.get('colspan', 1) > 1 or cell_info.get('rowspan', 1) > 1:
                ws.merge_cells(
                    start_row=c_row, 
                    start_column=current_col, 
                    end_row=c_row + cell_info.get('rowspan', 1) - 1, 
                    end_column=current_col + cell_info.get('colspan', 1) - 1
                )

    # Data row
    ws.cell(row=6, column=2, value=100)
    ws.cell(row=6, column=3, value=50)

    # _autosize_excel logic
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            if cell.value:
                try:
                    lines = str(cell.value).split("\n")
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
                except:
                    pass
        
        if col_letter:
            adjusted_width = min(max(max_length + 2, 12), 60)
            ws.column_dimensions[col_letter].width = adjusted_width
            
    print("Col B width:", ws.column_dimensions['B'].width)
    print("Col C width:", ws.column_dimensions['C'].width)

simulate()
