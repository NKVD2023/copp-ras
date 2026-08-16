import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active

# Simulate a horizontally merged cell above it (row 2, cols A, B, C)
ws.cell(row=2, column=1, value="Top Header Spanning A-C")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

# Simulate a vertically merged leaf cell in column B
ws.cell(row=3, column=2, value="Leaf header in Col B that is very long")
ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)

def _autosize_excel(ws, min_col_width=12, max_col_width=60, default_row_height=15):
    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if col_letter is None:
                col_letter = get_column_letter(cell.column)
            if cell.value:
                print(f"Row {cell.row} Col {cell.column} value={cell.value}")
                try:
                    lines = str(cell.value).split("\n")
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
                except:
                    pass
        
        if col_letter:
            adjusted_width = min(max(max_length + 2, min_col_width), max_col_width)
            ws.column_dimensions[col_letter].width = adjusted_width

_autosize_excel(ws)
print("Col A width:", ws.column_dimensions['A'].width)
print("Col B width:", ws.column_dimensions['B'].width)
print("Col C width:", ws.column_dimensions['C'].width)

